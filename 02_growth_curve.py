"""
02_growth_curve.py
==================
#env = mAIcrobe
OD600 growth curve analysis for CRISPRi mutants.

Input:  3 biological replicate kinetic OD600 CSV files (96-well, 120 time points, 10 min apart)
        1 plate map CSV (8 rows × 12 cols, semicolon-delimited, no header)
Output: Per-mutant growth curve plots (raw + mean±std + 4PL fit)
        Overview grid of all genes (colored by gene/pathway)
        Growth parameter table (CSV + Excel sorted by EC50)
"""

from __future__ import annotations

import colorsys
import re
import warnings
from pathlib import Path

import matplotlib.colors as mcolors
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.optimize import OptimizeWarning, curve_fit
from tqdm import tqdm

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

class Config:
    DATA_DIR = Path(r"D:\2026_04_10 CRISPRi Growth Curves")
    PLATE_MAP = DATA_DIR / "CRISPRI Reference Plate Map.csv"
    BIO_REP_FILES = [
        DATA_DIR / "260401_OD600 20h kinetic(FK)_2026_03_31_CRISPRi Kinetics_#1.csv",
        DATA_DIR / "260402_OD600 20h kinetic(FK)_2026_04_01_CRISPRi Kinetics_#2.csv",
        DATA_DIR / "260413_OD600 20h kinetic(FK)_2026_04_10_CRISPRi Kinetics_#3.csv",
    ]
    OUTPUT_DIR    = DATA_DIR / "growth_curve_analysis"
    PER_MUTANT_DIR = OUTPUT_DIR / "per_mutant"
    RAW_DIR        = PER_MUTANT_DIR / "raw"
    SUMMARY_DIR    = PER_MUTANT_DIR / "summary"
    DPI = 300

    # Base gene colors (one per gene, from 03_morphological_map.py palette)
    GENE_COLORS: dict[str, str] = {
        # Cell wall synthesis
        "mrcA": "#E57373", "mrcB": "#EF5350", "mrdA": "#F06292",
        "ftsI": "#EC407A", "murA": "#FFB74D", "murC": "#FFA726",
        # LPS synthesis
        "lpxA": "#4DB6AC", "lpxC": "#26A69A", "lptA": "#4DD0E1",
        "lptC": "#26C6DA", "msbA": "#80DEEA",
        # DNA metabolism
        "gyrA": "#5C6BC0", "gyrB": "#3F51B5", "parC": "#7986CB",
        "parE": "#9FA8DA", "dnaE": "#9575CD", "dnaB": "#B39DDB",
        # Transcription & translation
        "rpoA": "#81C784", "rpoB": "#66BB6A", "rpsA": "#FFF176",
        "rpsL": "#FFEE58", "rplA": "#FFD54F", "rplC": "#FFCA28",
        # Metabolism & export
        "folA": "#AED581", "folP": "#9CCC65", "secY": "#80CBC4", "secA": "#4DB6AC",
        # Cell division
        "ftsZ": "#F06292",
        # Controls
        "WT": "#000000", "NC": "#BDBDBD",
    }

    # Pathway grouping
    GENE_PATHWAY: dict[str, str] = {
        "mrcA": "Cell wall", "mrcB": "Cell wall", "mrdA": "Cell wall",
        "ftsI": "Cell wall", "murA": "Cell wall", "murC": "Cell wall",
        "lpxA": "LPS", "lpxC": "LPS", "lptA": "LPS",
        "lptC": "LPS", "msbA": "LPS",
        "gyrA": "DNA", "gyrB": "DNA", "parC": "DNA",
        "parE": "DNA", "dnaE": "DNA", "dnaB": "DNA",
        "rpoA": "Transcription", "rpoB": "Transcription",
        "rpsA": "Translation", "rpsL": "Translation",
        "rplA": "Translation", "rplC": "Translation",
        "folA": "Metabolism", "folP": "Metabolism",
        "secY": "Secretion", "secA": "Secretion",
        "ftsZ": "Cell division",
        "WT": "Control", "NC": "Control",
    }

    WT_COLOR  = "#888888"
    FIT_COLOR = "#1A1A1A"

    # 4PL fit bounds: (bottom, top, ec50_h, hill)
    FIT_BOUNDS_LOW  = (0.0, 0.0,  0.01, 0.1)
    FIT_BOUNDS_HIGH = (1.5, 2.0, 25.0, 20.0)

    # Outlier exclusion: individual curves whose EC50 deviates more than
    # this many standard deviations from the group mean are excluded before
    # computing the mean curve and its final 4PL fit.
    OUTLIER_N_STD = 2.0


# ---------------------------------------------------------------------------
# Color helpers
# ---------------------------------------------------------------------------

def _shade_color(hex_color: str, lightness_delta: float) -> str:
    """Shift the HLS lightness of a hex color by delta, return new hex."""
    r, g, b = mcolors.to_rgb(hex_color)
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    l_new = float(np.clip(l + lightness_delta, 0.08, 0.92))
    r2, g2, b2 = colorsys.hls_to_rgb(h, l_new, s)
    return mcolors.to_hex((r2, g2, b2))


def guide_colors(gene: str, guides: list[str]) -> list[str]:
    """
    Return one shade per guide, evenly spaced from lighter to darker.
    Works for any number of guides (3 for mutants, 6 for WT/NC controls).
    """
    base = Config.GENE_COLORS.get(gene, "#999999")
    n = len(guides)
    if n == 1:
        return [base]
    deltas = np.linspace(+0.22, -0.22, n)
    return [_shade_color(base, float(d)) for d in deltas]


def hex_to_openpyxl_rgb(hex_color: str) -> str:
    """Convert '#RRGGBB' to 'FFRRGGBB' (openpyxl ARGB format)."""
    h = hex_color.lstrip("#")
    return "FF" + h.upper()


# ---------------------------------------------------------------------------
# Plate map
# ---------------------------------------------------------------------------

def read_plate_map(path: Path) -> pd.DataFrame:
    """Read 8×12 plate map (semicolon-delimited, no header)."""
    pm = pd.read_csv(path, sep=";", header=None)
    pm = pm.map(lambda x: str(x).strip() if pd.notna(x) else x)
    return pm


def build_well_label_dict(plate_map: pd.DataFrame) -> dict[str, str]:
    """Pre-compute {well: label} for all 96 wells."""
    mapping = {}
    for r, row_letter in enumerate("ABCDEFGH"):
        for c in range(1, 13):
            well = f"{row_letter}{c:02d}"
            mapping[well] = str(plate_map.iloc[r, c - 1]).strip()
    return mapping


# ---------------------------------------------------------------------------
# Gene / guide parsing
# ---------------------------------------------------------------------------

def parse_gene_guide(label: str) -> tuple[str, str | None]:
    """
    'mrcA_1'  -> ('mrcA', '1')   guides 1–3
    'WT NC_1' -> ('WT',   '1')   guides 1–6
    'NC_3'    -> ('NC',   '3')   guides 1–6
    """
    label = label.strip()
    parts = label.rsplit("_", 1)
    guide = parts[1] if len(parts) == 2 and parts[1].isdigit() else None
    if label.startswith("WT"):
        return "WT", guide         # 'WT NC_1' → ('WT', '1'), guides 1–6
    if label.startswith("NC"):
        return "NC", guide         # 'NC_3'    → ('NC', '3'), guides 1–6
    if guide is not None:
        return parts[0], guide     # 'mrcA_2'  → ('mrcA', '2'), guides 1–3
    return label, None


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def read_od600_file(path: Path, replicate_idx: int) -> pd.DataFrame:
    """
    Parse one kinetic OD600 CSV from the plate reader.

    File structure:
        8 metadata lines
        Line: 'Time [s]:'
        Line: <empty>;<t0>;<t1>;...   (time values in seconds)
        Lines: <well>;<od0>;<od1>;... (one per well, A01-H12)

    Returns long-form DataFrame: [well, time_h, OD600, replicate]
    """
    time_hours: list[float] | None = None
    next_is_time = False
    records: list[dict] = []

    well_pattern = re.compile(r"^[A-H]\d{2}$")

    with open(path, "r", encoding="utf-8") as fh:
        for raw_line in fh:
            line = raw_line.rstrip("\n")
            parts = [p.strip() for p in line.split(";")]
            first = parts[0]

            if first == "Time [s]:":
                next_is_time = True
                continue

            if next_is_time:
                time_s = [float(p) for p in parts if p]
                time_hours = [t / 3600.0 for t in time_s]
                next_is_time = False
                continue

            if well_pattern.match(first):
                od_values = [float(p) for p in parts[1:] if p]
                if time_hours is None:
                    raise ValueError(f"Time row not found before well data in {path.name}")
                for t, od in zip(time_hours, od_values):
                    records.append({
                        "well": first,
                        "time_h": t,
                        "OD600": od,
                        "replicate": replicate_idx,
                    })

    return pd.DataFrame(records)


def load_all_data() -> pd.DataFrame:
    """Load plate map + all 3 bio replicate files; return annotated long-form DataFrame."""
    plate_map = read_plate_map(Config.PLATE_MAP)
    well_labels = build_well_label_dict(plate_map)

    dfs = []
    for idx, path in enumerate(Config.BIO_REP_FILES, start=1):
        print(f"  Loading replicate {idx}: {path.name}")
        df = read_od600_file(path, replicate_idx=idx)
        dfs.append(df)

    data = pd.concat(dfs, ignore_index=True)

    data["label"] = data["well"].map(well_labels)
    data[["gene", "guide"]] = data["label"].apply(
        lambda lbl: pd.Series(parse_gene_guide(lbl))
    )

    return data


# ---------------------------------------------------------------------------
# 4PL sigmoidal fit
# ---------------------------------------------------------------------------

def _four_pl(x: np.ndarray, bottom: float, top: float, ec50: float, hill: float) -> np.ndarray:
    """Four-parameter logistic: y = Bottom + (Top-Bottom) / (1 + (EC50/x)^Hill)"""
    with np.errstate(divide="ignore", invalid="ignore"):
        return bottom + (top - bottom) / (1.0 + (ec50 / x) ** hill)


def fit_4pl(time_h: np.ndarray, od600_mean: np.ndarray) -> dict | None:
    """
    Fit a 4PL curve to the mean OD600 time series.
    Returns dict with keys: bottom, top, ec50_h, hill, fitted_curve
    Returns None if fit fails.
    """
    mask = time_h > 0
    t = time_h[mask]
    y = od600_mean[mask]

    bottom_init = float(np.percentile(y, 5))
    top_init    = float(np.percentile(y, 95))
    mid = (bottom_init + top_init) / 2.0
    ec50_init = float(t[np.argmin(np.abs(y - mid))]) if len(t) else 5.0
    p0 = [bottom_init, top_init, max(ec50_init, 0.1), 2.0]

    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", OptimizeWarning)
            popt, _ = curve_fit(
                _four_pl, t, y,
                p0=p0,
                bounds=(Config.FIT_BOUNDS_LOW, Config.FIT_BOUNDS_HIGH),
                maxfev=10_000,
            )
        return {
            "bottom":       popt[0],
            "top":          popt[1],
            "ec50_h":       popt[2],
            "hill":         popt[3],
            "fitted_curve": _four_pl(time_h, *popt),
        }
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Per-curve fitting & outlier detection
# ---------------------------------------------------------------------------

def fit_all_individual_curves(gene_data: pd.DataFrame) -> pd.DataFrame:
    """
    Fit 4PL independently to every individual curve (guide × replicate).
    Returns DataFrame with columns: guide, replicate, ec50_h, fit_ok
    """
    rows = []
    for (guide, rep), grp in gene_data.groupby(["guide", "replicate"], sort=True):
        grp = grp.sort_values("time_h")
        fit = fit_4pl(grp["time_h"].values, grp["OD600"].values)
        rows.append({
            "guide":     guide,
            "replicate": rep,
            "ec50_h":    fit["ec50_h"] if fit else np.nan,
            "fit_ok":    fit is not None,
        })
    return pd.DataFrame(rows)


def flag_outliers(
    individual_fits: pd.DataFrame,
    n_std: float = Config.OUTLIER_N_STD,
) -> tuple[pd.DataFrame, float, float]:
    """
    Mark curves as outliers when their EC50 deviates > n_std × std from the mean.
    Returns (annotated_df, mean_ec50, std_ec50).
    Failed fits (ec50=NaN) are always flagged as outliers.
    Falls back to no exclusion if fewer than 2 valid fits exist.
    """
    valid = individual_fits["ec50_h"].dropna()
    if len(valid) < 2:
        individual_fits = individual_fits.copy()
        individual_fits["is_outlier"] = ~individual_fits["fit_ok"]
        return individual_fits, float(valid.mean()) if len(valid) else np.nan, np.nan

    mean_ec50 = float(valid.mean())
    std_ec50  = float(valid.std())

    is_outlier = (
        individual_fits["ec50_h"].isna()
        | ((individual_fits["ec50_h"] - mean_ec50).abs() > n_std * std_ec50)
    )
    individual_fits = individual_fits.copy()
    individual_fits["is_outlier"] = is_outlier
    return individual_fits, mean_ec50, std_ec50


# ---------------------------------------------------------------------------
# Plotting helpers
# ---------------------------------------------------------------------------

def _style_ax(ax: plt.Axes) -> None:
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.xaxis.set_minor_locator(ticker.AutoMinorLocator(2))
    ax.yaxis.set_minor_locator(ticker.AutoMinorLocator(2))
    ax.tick_params(which="both", direction="in")


def _get_mean_std(data: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    """Compute mean ± std of OD600 grouped by time_h + group_cols."""
    return (
        data.groupby(["time_h"] + group_cols, sort=True)["OD600"]
        .agg(mean="mean", std="std")
        .reset_index()
    )


# ---------------------------------------------------------------------------
# Per-mutant plot
# ---------------------------------------------------------------------------

def _prepare_gene_context(
    gene: str,
    gene_data: pd.DataFrame,
    wt_data: pd.DataFrame,
) -> dict:
    """
    Shared pre-computation for both raw and summary plots:
      - WT NC reference stats
      - Per-curve 4PL fits + outlier flagging
      - Retained vs outlier key sets
      - Guide colors
      - Single 4PL fit on mean of retained curves
    Returns a context dict consumed by the two plot functions.
    """
    wt_stats = _get_mean_std(wt_data, [])

    ind_fits = fit_all_individual_curves(gene_data)
    ind_fits, _, _ = flag_outliers(ind_fits)
    n_outliers = int(ind_fits["is_outlier"].sum())

    retained_keys = set(zip(
        ind_fits.loc[~ind_fits["is_outlier"], "guide"],
        ind_fits.loc[~ind_fits["is_outlier"], "replicate"],
    ))
    outlier_keys = set(zip(
        ind_fits.loc[ind_fits["is_outlier"], "guide"],
        ind_fits.loc[ind_fits["is_outlier"], "replicate"],
    ))

    # EC50 std from retained individual fits only
    retained_ec50_vals = ind_fits.loc[~ind_fits["is_outlier"], "ec50_h"].dropna()
    ec50_ind_std = float(retained_ec50_vals.std()) if len(retained_ec50_vals) > 1 else np.nan

    retained_mask = np.array(
        [(g, r) in retained_keys for g, r in zip(gene_data["guide"], gene_data["replicate"])],
        dtype=bool,
    )
    retained_data: pd.DataFrame = gene_data.loc[retained_mask]
    if retained_data.empty:
        retained_data = gene_data.copy()

    retained_overall = _get_mean_std(retained_data, [])
    fit_result = fit_4pl(retained_overall["time_h"].to_numpy(), retained_overall["mean"].to_numpy())

    guides = sorted(gene_data["guide"].dropna().unique(), key=lambda x: int(x))
    colors = guide_colors(gene, guides)

    guide_stats_list: list[tuple] = []
    for guide, color in zip(guides, colors):
        g_ret: pd.DataFrame = retained_data.loc[retained_data["guide"] == guide]
        if not g_ret.empty:
            guide_stats_list.append((guide, color, _get_mean_std(g_ret, [])))

    return {
        "gene":             gene,
        "gene_data":        gene_data,
        "wt_stats":         wt_stats,
        "retained_data":    retained_data,
        "outlier_keys":     outlier_keys,
        "n_outliers":       n_outliers,
        "fit_result":       fit_result,
        "ec50_ind_std":     ec50_ind_std,
        "guides":           guides,
        "colors":           colors,
        "guide_stats_list": guide_stats_list,
        "time_pts":         np.sort(gene_data["time_h"].unique()),
    }


def plot_raw(ctx: dict, raw_dir: Path, y_lim: tuple[float, float] = None) -> None:
    """
    Raw data plot: every individual curve, outliers as grey dashed lines.
    No means, no fit — data only.
    """
    gene        = ctx["gene"]
    gene_data   = ctx["gene_data"]
    wt_stats    = ctx["wt_stats"]
    outlier_keys = ctx["outlier_keys"]
    guides      = ctx["guides"]
    colors      = ctx["colors"]
    n_outliers  = ctx["n_outliers"]

    fig, ax = plt.subplots(figsize=(10, 5))
    outlier_note = f"  ({n_outliers} outlier{'s' if n_outliers != 1 else ''} excluded)" if n_outliers else ""
    ax.set_title(f"{gene}  —  raw curves{outlier_note}", fontsize=13, fontweight="bold")

    ax.plot(
        wt_stats["time_h"], wt_stats["mean"].to_numpy(),
        color=Config.WT_COLOR, linewidth=1.5, linestyle="--", label="WT NC", zorder=2,
    )
    ax.set_xlim(left=0)
    if y_lim is not None:
        ax.set_ylim(y_lim)

    _outlier_labeled = False
    for guide, color in zip(guides, colors):
        g_data: pd.DataFrame = gene_data.loc[gene_data["guide"] == guide]
        reps = sorted(g_data["replicate"].unique())
        for rep_idx in reps:
            rep_data = g_data.loc[g_data["replicate"] == rep_idx].sort_values("time_h")
            if (guide, rep_idx) in outlier_keys:
                lbl = "Outlier (excluded)" if not _outlier_labeled else None
                _outlier_labeled = True
                ax.plot(rep_data["time_h"], rep_data["OD600"],
                        color="#BBBBBB", alpha=0.7, linewidth=0.8,
                        linestyle="--", zorder=1, label=lbl)
            else:
                ax.plot(rep_data["time_h"], rep_data["OD600"],
                        color=color, alpha=0.35, linewidth=0.9, zorder=1,
                        label=f"Guide {guide}" if rep_idx == reps[0] else None)

    ax.set_xlabel("Time (h)")
    ax.set_ylabel("OD600")
    ax.legend(fontsize=8, frameon=False)
    _style_ax(ax)
    plt.tight_layout()
    fig.savefig(str(raw_dir / f"{gene}_raw.png"), dpi=Config.DPI, bbox_inches="tight")
    plt.close(fig)


def plot_summary(ctx: dict, summary_dir: Path, y_lim: tuple[float, float] = None) -> None:
    """
    Summary plot: per-guide mean ± std ribbons (retained curves only)
    + WT NC reference + single 4PL fit on overall retained mean.
    """
    gene             = ctx["gene"]
    wt_stats         = ctx["wt_stats"]
    guide_stats_list = ctx["guide_stats_list"]
    fit_result       = ctx["fit_result"]
    ec50_ind_std     = ctx["ec50_ind_std"]
    time_pts         = ctx["time_pts"]
    n_outliers       = ctx["n_outliers"]

    wt_mean = wt_stats["mean"].to_numpy()
    wt_std  = wt_stats["std"].fillna(0).to_numpy()

    fig, ax = plt.subplots(figsize=(10, 5))
    outlier_note = f"  ({n_outliers} outlier{'s' if n_outliers != 1 else ''} excluded)" if n_outliers else ""
    ax.set_title(f"{gene}  —  summary{outlier_note}", fontsize=13, fontweight="bold")
    ax.set_xlim(left=0)
    if y_lim is not None:
        ax.set_ylim(y_lim)

    # WT NC ribbon + mean
    ax.fill_between(wt_stats["time_h"], wt_mean - wt_std, wt_mean + wt_std,
                    color=Config.WT_COLOR, alpha=0.12, zorder=1)
    ax.plot(wt_stats["time_h"], wt_mean,
            color=Config.WT_COLOR, linewidth=1.5, linestyle="--", label="WT NC", zorder=2)

    # Per-guide mean ± std
    for guide, color, guide_stats in guide_stats_list:
        t_vals    = guide_stats["time_h"].to_numpy()
        mean_vals = guide_stats["mean"].to_numpy()
        std_vals  = guide_stats["std"].fillna(0).to_numpy()
        ax.fill_between(t_vals, mean_vals - std_vals, mean_vals + std_vals,
                        color=color, alpha=0.2, zorder=2)
        ax.plot(t_vals, mean_vals, color=color, linewidth=2.0,
                label=f"Guide {guide}", zorder=3)

    # Single 4PL fit
    if fit_result is not None:
        ec50 = fit_result["ec50_h"]
        if np.isfinite(ec50_ind_std):
            ec50_label = f"4PL fit  (EC50 = {ec50:.1f} ± {ec50_ind_std:.1f} h)"
        else:
            ec50_label = f"4PL fit  (EC50 = {ec50:.1f} h)"
        ax.plot(time_pts, fit_result["fitted_curve"],
                color=Config.FIT_COLOR, linewidth=2.2,
                label=ec50_label, zorder=4)
        ax.axvline(ec50, color=Config.FIT_COLOR, linewidth=0.8, linestyle=":", alpha=0.7)

    ax.set_xlabel("Time (h)")
    ax.set_ylabel("OD600")
    ax.legend(fontsize=8, frameon=False)
    _style_ax(ax)
    plt.tight_layout()
    fig.savefig(str(summary_dir / f"{gene}_summary.png"), dpi=Config.DPI, bbox_inches="tight")
    plt.close(fig)


# ---------------------------------------------------------------------------
# Overview grid
# ---------------------------------------------------------------------------

def plot_overview_grid(
    all_data: pd.DataFrame,
    wt_data: pd.DataFrame,
    output_dir: Path,
) -> None:
    """7×4 grid of all 28 mutant mean curves, colored by gene, vs WT NC reference."""
    mutant_genes = sorted(
        g for g in all_data["gene"].unique()
        if g not in ("WT", "NC")
    )

    n_cols = 4
    n_rows = int(np.ceil(len(mutant_genes) / n_cols))

    wt_stats = _get_mean_std(wt_data, [])

    # Fixed 16:9 landscape canvas
    fig, axes = plt.subplots(
        n_rows, n_cols,
        figsize=(28, 15.75),
        sharex=True,
    )
    axes_flat = axes.flatten()

    for ax, gene in zip(axes_flat, mutant_genes):
        gene_data  = all_data[all_data["gene"] == gene]
        gene_stats = _get_mean_std(gene_data, [])
        color = Config.GENE_COLORS.get(gene, "#999999")

        ax.plot(
            wt_stats["time_h"], wt_stats["mean"],
            color=Config.WT_COLOR, linewidth=1.0, linestyle="--",
        )
        ax.fill_between(
            gene_stats["time_h"],
            gene_stats["mean"] - gene_stats["std"].fillna(0),
            gene_stats["mean"] + gene_stats["std"].fillna(0),
            color=color, alpha=0.2,
        )
        ax.plot(gene_stats["time_h"], gene_stats["mean"], color=color, linewidth=1.8)

        ax.set_title(gene, fontsize=10, fontweight="bold", color=color)
        _style_ax(ax)
        ax.tick_params(labelsize=7)

    for ax in axes_flat[len(mutant_genes):]:
        ax.set_visible(False)

    fig.supxlabel("Time (h)", fontsize=11)
    fig.supylabel("OD600", fontsize=11)
    fig.suptitle("CRISPRi Growth Curves — Overview", fontsize=14, fontweight="bold", y=1.01)
    plt.tight_layout()

    out_path = output_dir / "overview_grid.png"
    fig.savefig(out_path, dpi=Config.DPI, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {out_path.name}")


# ---------------------------------------------------------------------------
# Growth parameter table
# ---------------------------------------------------------------------------

def compute_growth_parameters(all_data: pd.DataFrame) -> pd.DataFrame:
    """
    For each gene, fit 4PL to the overall mean curve.
    Returns DataFrame sorted by ec50_h with columns:
        gene, pathway, bottom, top, ec50_h, hill_slope, max_OD600_mean, max_OD600_std, n_curves
    """
    rows = []
    for gene in sorted(all_data["gene"].unique()):
        gene_data: pd.DataFrame = pd.DataFrame(all_data.loc[all_data["gene"] == gene])

        # Outlier exclusion — mirrors _prepare_gene_context exactly
        ind_fits = fit_all_individual_curves(gene_data)
        ind_fits, _, _ = flag_outliers(ind_fits)
        retained_keys = set(zip(
            ind_fits.loc[~ind_fits["is_outlier"], "guide"],
            ind_fits.loc[~ind_fits["is_outlier"], "replicate"],
        ))
        retained_mask = np.array(
            [(g, r) in retained_keys
             for g, r in zip(gene_data["guide"], gene_data["replicate"])],
            dtype=bool,
        )
        retained_data = gene_data.loc[retained_mask]
        if retained_data.empty:
            retained_data = gene_data.copy()

        retained_ec50 = ind_fits.loc[~ind_fits["is_outlier"], "ec50_h"].dropna()
        ec50_std = round(float(retained_ec50.std()), 3) if len(retained_ec50) > 1 else np.nan

        stats = _get_mean_std(retained_data, [])
        t = stats["time_h"].to_numpy()
        mean_od = stats["mean"].to_numpy()

        fit = fit_4pl(t, mean_od)
        n_curves = gene_data["replicate"].nunique() * gene_data["guide"].dropna().nunique()

        row: dict = {
            "gene":           gene,
            "pathway":        Config.GENE_PATHWAY.get(gene, "Other"),
            "n_curves":       n_curves,
            "max_OD600_mean": round(float(np.nanmax(mean_od)), 4),
            "max_OD600_std":  round(float(stats["std"].max(skipna=True)), 4),
        }
        if fit:
            b, t_, ec50, hill = fit["bottom"], fit["top"], fit["ec50_h"], fit["hill"]
            # Max growth rate = slope at inflection point of 4PL (OD/h)
            max_gr = (t_ - b) * hill / (4.0 * ec50) if ec50 > 0 else np.nan
            row.update({
                "bottom":          round(b, 4),
                "top":             round(t_, 4),
                "ec50_h":          round(ec50, 3),
                "ec50_std":        ec50_std,
                "hill_slope":      round(hill, 3),
                "max_growth_rate": round(max_gr, 4),
            })
        else:
            row.update({
                "bottom": np.nan, "top": np.nan, "ec50_h": np.nan,
                "ec50_std": ec50_std,
                "hill_slope": np.nan, "max_growth_rate": np.nan,
            })

        rows.append(row)

    df = pd.DataFrame(rows, columns=[
        "gene", "pathway", "bottom", "top", "ec50_h", "ec50_std", "hill_slope",
        "max_growth_rate", "max_OD600_mean", "max_OD600_std", "n_curves",
    ])
    return df.sort_values("ec50_h").reset_index(drop=True)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def save_excel(params: pd.DataFrame, output_dir: Path) -> None:
    """
    Save growth parameters as Excel, sorted by EC50.
    Rows are color-coded by pathway using gene base colors.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Growth Parameters"

    header = list(params.columns)
    ws.append(header)

    # Bold header row
    for col_idx, _ in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data rows with pathway-based row fill
    for row_idx, row in params.iterrows():
        ws.append(list(row))
        excel_row = row_idx + 2   # 1-indexed + 1 header

        gene  = row["gene"]
        color = Config.GENE_COLORS.get(gene, "#FFFFFF")
        # Lighten the gene color significantly for readability as a background
        bg = _shade_color(color, +0.30) if color != "#000000" else "#EEEEEE"
        fill = PatternFill(start_color=hex_to_openpyxl_rgb(bg),
                           end_color=hex_to_openpyxl_rgb(bg),
                           fill_type="solid")
        for col_idx in range(1, len(header) + 1):
            ws.cell(row=excel_row, column=col_idx).fill = fill
            ws.cell(row=excel_row, column=col_idx).alignment = Alignment(horizontal="center")

    # Auto-fit column widths
    for col_idx, col_name in enumerate(header, start=1):
        max_len = max(len(str(col_name)), 10)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    out_path = output_dir / "growth_parameters_sorted_EC50.xlsx"
    wb.save(out_path)
    print(f"  Saved: {out_path.name}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("=== OD600 Growth Curve Analysis ===")

    Config.RAW_DIR.mkdir(parents=True, exist_ok=True)
    Config.SUMMARY_DIR.mkdir(parents=True, exist_ok=True)

    # 1 — Load
    print("\n[1/5] Loading data...")
    all_data = load_all_data()
    print(f"  Loaded {len(all_data):,} measurements | "
          f"{all_data['gene'].nunique()} genes | "
          f"{all_data['replicate'].nunique()} bio replicates")

    wt_data: pd.DataFrame = pd.DataFrame(all_data.loc[all_data["gene"] == "WT"])

    # 2 — Per-mutant plots (raw + summary in separate subfolders)
    genes = sorted(all_data["gene"].unique())
    print(f"\n[2/5] Plotting {len(genes)} gene(s)  ->  raw/ and summary/...")
    
    global_min = 0
    global_max = 1.25
    
    for gene in tqdm(genes):
        gene_data: pd.DataFrame = pd.DataFrame(all_data.loc[all_data["gene"] == gene])
        ctx = _prepare_gene_context(gene, gene_data, wt_data)
        plot_raw(ctx, Config.RAW_DIR, y_lim=(global_min, global_max))
        plot_summary(ctx, Config.SUMMARY_DIR, y_lim=(global_min, global_max))

    # 3 — Overview grid
    print("\n[3/5] Generating overview grid...")
    plot_overview_grid(all_data, wt_data, Config.OUTPUT_DIR)

    # 4 — Growth parameters
    print("\n[4/5] Computing growth parameters...")
    params = compute_growth_parameters(all_data)
    csv_path = Config.OUTPUT_DIR / "growth_parameters.csv"
    params.to_csv(csv_path, index=False)
    print(f"  Saved: {csv_path.name}")
    print(params.to_string(index=False))

    # 5 — Excel export sorted by EC50
    print("\n[5/5] Saving Excel (sorted by EC50)...")
    save_excel(params, Config.OUTPUT_DIR)

    print("\nDone.")


if __name__ == "__main__":
    main()
